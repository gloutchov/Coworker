import csv
import json

from pathlib import Path
from typing import List, Dict, Any

from bs4 import BeautifulSoup
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook

from modules.audio.transcriber import transcribe_audio_file

from config import (
    RAG_CHUNK_MAX_CHARS, RAG_CHUNK_OVERLAP, RAG_TOP_K_DEFAULT,
    RAG_CHUNK_MAX_LINES, RAG_CHUNK_OVERLAP_LINES,
    RAG_MIN_KEYWORD_LEN,
    # nuovi import per whisper
    WHISPER_MODEL_PATH, WHISPER_DEVICE, WHISPER_COMPUTE_TYPE,
)

from modules.db.rag import (
    clear_chunks,
    insert_chunks,
    get_all_chunks,
    clear_documents,
    upsert_documents,
    delete_document_from_rag,
)
from modules.db.chats import (
    clear_chat_chunks_for_chat,
    insert_chat_chunks,
    get_all_chat_chunks,
)

from llm_client import get_embedding
from modules.config.preferences import get_excel_limits
from modules.config.preferences import get_docs_dir

# Estensioni audio che vogliamo trattare come "documenti da trascrivere"
AUDIO_SUFFIXES = {".wav", ".mp3", ".m4a", ".aac", ".ogg", ".flac", ".webm", ".opus"}


def extract_visible_text_from_html(html: str) -> str:
    """
    Estrae il testo "visibile" da una stringa HTML.
    Replica la funzione che hai già in server.py.
    """
    soup = BeautifulSoup(html, "html.parser")

    for tag in soup(["script", "style", "noscript", "iframe"]):
        tag.decompose()

    root = soup.body if soup.body is not None else soup

    raw_text = root.get_text(separator="\n", strip=True)

    lines = [line.strip() for line in raw_text.splitlines()]
    cleaned_lines: List[str] = []
    for line in lines:
        if not line:
            if cleaned_lines and cleaned_lines[-1] != "":
                cleaned_lines.append("")
        else:
            cleaned_lines.append(line)

    return "\n".join(cleaned_lines).strip()


def extract_text_from_file(path: Path) -> str:
    """
    Restituisce il testo estratto dal file indicato.
    Supporta: .txt, .md, .pdf, .docx, .xlsx, .csv, .js, .py, .html, .htm, .css, e file audio.
    """
    from PyPDF2.errors import PdfReadError  # opzionale, se vuoi distinguere

    suffix = path.suffix.lower()

    try:
        # --- Gestione file audio tramite il nuovo modulo centralizzato ---
        if suffix in AUDIO_SUFFIXES:
            result = transcribe_audio_file(str(path))
            if not result:
                print(f"[RAG] Nessun testo estratto dal file audio: {path.name}")
                return ""
            text = (result.get("text") or "").strip()
            if not text:
                print(f"[RAG] Nessun testo estratto dal file audio: {path.name}")
                return ""
            # Aggiungo un header per riconoscere che il testo viene da un audio
            header = f"[TRASCRIZIONE AUDIO] File: {path.name}\n"
            return header + text

        if suffix in [".txt", ".md", ".js", ".py", ".css"]:
            return path.read_text(encoding="utf-8", errors="ignore")

        if suffix in [".html", ".htm"]:
            html = path.read_text(encoding="utf-8", errors="ignore")
            return extract_visible_text_from_html(html)

        if suffix == ".pdf":
            text_parts: List[str] = []
            reader = PdfReader(str(path))
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text_parts.append(page_text)
            return "\n".join(text_parts)

        if suffix == ".docx":
            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)

        if suffix == ".xlsx":
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            text_parts: List[str] = []
            max_rows, max_cols = get_excel_limits()

            for ws in wb.worksheets:
                text_parts.append(f"=== SHEET: {ws.title} ===")

                header = None

                for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if r_idx > max_rows:
                        text_parts.append("...[righe troncate]...")
                        break

                    # normalizza celle in stringhe (SOLO per la riga corrente)
                    raw_cells = list(row[:max_cols])
                    cells = [
                        (str(c).strip() if c is not None else "")
                        for c in raw_cells
                    ]

                    if r_idx == 1:
                        # prima riga: intestazioni
                        header = cells
                        header_str = " | ".join(h for h in header if h)
                        text_parts.append(f"HEADER: {header_str}")
                        continue

                    # righe dati
                    if header:
                        pairs = []
                        for h, c in zip(header, cells):
                            h = (h or "").strip()
                            c = (c or "").strip()
                            if c:
                                if h:
                                    pairs.append(f"{h}: {c}")
                                else:
                                    pairs.append(c)
                        if pairs:
                            text_parts.append(" | ".join(pairs))
                    else:
                        # fallback: nessuna intestazione valida
                        cells_str = [c for c in cells if c]
                        if cells_str:
                            text_parts.append(" | ".join(cells_str))

            wb.close()
            return "\n".join(text_parts)


        if suffix == ".csv":
            text_parts: List[str] = []

            max_rows, max_cols = get_excel_limits()

            with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel

                reader = csv.reader(f, dialect=dialect)

                header = None
                for r_idx, row in enumerate(reader, start=1):
                    if r_idx == 1:
                        header = row[:max_cols]
                        header_str = " | ".join((h or "").strip() for h in header)
                        text_parts.append(f"HEADER: {header_str}")
                        continue

                    if r_idx > max_rows:
                        text_parts.append("...[righe troncate]...")
                        break

                    cells = row[:max_cols]

                    if header:
                        pairs = []
                        for h, c in zip(header, cells):
                            h = (h or "").strip()
                            c = (c or "").strip()
                            if c:
                                if h:
                                    pairs.append(f"{h}: {c}")
                                else:
                                    pairs.append(c)
                        if pairs:
                            text_parts.append(" | ".join(pairs))
                    else:
                        cells_str = [(c or "").strip()
                                     for c in cells
                                     if c not in (None, "")]
                        if cells_str:
                            text_parts.append(" | ".join(cells_str))

            return "\n".join(text_parts)

    except Exception:
        return ""

    return ""


def split_into_chunks(text: str, max_chars: int = 800, overlap: int = 200) -> list[str]:
    """
    Divide il testo in chunk basati su gruppi di righe,
    con un limite massimo di caratteri per chunk.

    - Raggruppa ~5-6 righe per chunk, con una piccola sovrapposizione in righe.
    - Se un gruppo è troppo lungo (> max_chars), lo spezza comunque a caratteri
      come fallback.
    """

    # Parametri "line-based"
    MAX_LINES = 6          # quante righe per chunk (circa)
    OVERLAP_LINES = 1      # quante righe si sovrappongono tra un chunk e il successivo

    chunks: list[str] = []
    text = text.strip()
    if not text:
        return chunks

    # Spezza il testo in righe
    lines = text.splitlines()
    n = len(lines)
    i = 0

    while i < n:
        end = min(i + MAX_LINES, n)
        group = lines[i:end]

        # Togliamo righe completamente vuote all'inizio/fine del gruppo
        while group and not group[0].strip():
            group = group[1:]
        while group and not group[-1].strip():
            group = group[:-1]

        if not group:
            # gruppo vuoto, avanzo
            if end == n:
                break
            i = max(end - OVERLAP_LINES, end)
            continue

        chunk_text = "\n".join(group).strip()

        # Se il chunk in righe è entro il limite di caratteri, lo teniamo così
        if len(chunk_text) <= max_chars:
            chunks.append(chunk_text)
        else:
            # Fallback: spezzare questo chunk "a caratteri", con lo schema vecchio
            start_char = 0
            length = len(chunk_text)
            while start_char < length:
                sub_end = min(start_char + max_chars, length)
                sub_chunk = chunk_text[start_char:sub_end].strip()
                if sub_chunk:
                    chunks.append(sub_chunk)
                if sub_end == length:
                    break
                # riuso l'overlap (in caratteri) passato come parametro
                start_char = sub_end - overlap

        if end == n:
            break

        # Avanza con sovrapposizione in righe
        i = max(end - OVERLAP_LINES, end)

    return chunks


def cosine_similarity(a: list[float], b: list[float]) -> float:
    """
    Similarità coseno tra due vettori.
    Copiata dalla tua versione.
    """
    if not a or not b or len(a) != len(b):
        return 0.0

    dot = 0.0
    na = 0.0
    nb = 0.0

    for x, y in zip(a, b):
        dot += x * y
        na += x * x
        nb += y * y

    if na == 0.0 or nb == 0.0:
        return 0.0

    return dot / ((na ** 0.5) * (nb ** 0.5))


def _prepare_document_chunks(path: Path) -> tuple[list[Dict[str, Any]], Dict[str, Any]] | None:
    """
    Estrae testo, chunk e metadati per un singolo file.
    Restituisce (chunks, meta) oppure None se il file non è indicizzabile.
    """
    content = extract_text_from_file(path)
    if not content or not content.strip():
        return None

    chunks = split_into_chunks(content)
    chunk_records: list[Dict[str, Any]] = []

    for idx, chunk_text in enumerate(chunks):
        emb = get_embedding(chunk_text)
        if not emb:
            continue
        chunk_records.append(
            {
                "file": path.name,
                "chunk_index": idx,
                "text": chunk_text,
                "embedding": emb,
            }
        )

    if not chunk_records:
        return None

    stat = path.stat()
    meta = {
        "file": path.name,
        "mtime": stat.st_mtime,
        "size": stat.st_size,
    }
    return chunk_records, meta


def index_documents() -> tuple[int, int]:
    """
    Indicizza la directory DOCS_DIR pulendo l'intero indice esistente.
    Restituisce: (numero_file_leggibili, numero_chunk)
    """
    clear_chunks()
    clear_documents()

    file_count = 0
    all_chunks: List[Dict[str, Any]] = []
    docs_meta: List[Dict[str, Any]] = []

    docs_dir = get_docs_dir()
    for path in docs_dir.rglob("*"):
        if not path.is_file():
            continue

        prepared = _prepare_document_chunks(path)
        if not prepared:
            continue

        chunk_records, meta = prepared
        file_count += 1
        all_chunks.extend(chunk_records)
        docs_meta.append(meta)

    insert_chunks(all_chunks)
    upsert_documents(docs_meta)

    print(f"[REINDEX] File leggibili: {file_count}")
    print(f"[REINDEX] Chunk indicizzati: {len(all_chunks)}")

    return file_count, len(all_chunks)


def index_single_document(path: Path) -> tuple[int, int]:
    """
    Indicizza un singolo file senza toccare il resto dell'indice.
    Restituisce (1, numero_chunk) se il file è stato indicizzato, altrimenti (0, 0).
    """
    if not path.exists() or not path.is_file():
        return (0, 0)

    prepared = _prepare_document_chunks(path)
    if not prepared:
        return (0, 0)

    chunk_records, meta = prepared
    delete_document_from_rag(meta["file"])
    insert_chunks(chunk_records)
    upsert_documents([meta])
    print(f"[REINDEX FILE] {meta['file']} -> {len(chunk_records)} chunk")
    return (1, len(chunk_records))


def index_chat_session_for_rag(chat_id: int, title: str, content_json: str) -> None:
    """
    Indicizza una chat salvata nella tabella chat_chunks.

    content_json è una lista JSON di oggetti {question, answer}, come quella che salvi lato frontend.
    """
    try:
        messages = json.loads(content_json)
    except Exception:
        messages = []

    chunks: List[Dict[str, Any]] = []

    for idx, m in enumerate(messages):
        q = (m.get("question") or "").strip()
        a = (m.get("answer") or "").strip()

        if not q and not a:
            continue

        parts = []
        if q:
            parts.append(f"DOMANDA: {q}")
        if a:
            parts.append(f"RISPOSTA: {a}")

        chunk_text = "\n".join(parts).strip()
        if not chunk_text:
            continue

        emb = get_embedding(chunk_text)
        if not emb:
            continue

        chunks.append(
            {
                "chat_id": chat_id,
                "title": title,
                "msg_index": idx,
                "text": chunk_text,
                "embedding": emb,
            }
        )

    # Prima rimuoviamo eventuali chunk precedenti di quella chat, poi inseriamo i nuovi
    clear_chat_chunks_for_chat(chat_id)
    insert_chat_chunks(chunks)


import re

# get_all_chunks, get_embedding e RAG_TOP_K_DEFAULT sono già importati sopra
# cosine_similarity è definita in questo stesso file, quindi non serve importarla


def _keyword_tokens(text: str) -> List[str]:
    """
    Estrae token "significativi" (lunghezza >= RAG_MIN_KEYWORD_LEN) in minuscolo.
    Serve per il match lessicale domanda -> chunk.
    """
    tokens = re.findall(r"\w+", text.lower())
    return [t for t in tokens if len(t) >= RAG_MIN_KEYWORD_LEN]


def get_relevant_chunks(question: str,
                        top_k: int = RAG_TOP_K_DEFAULT) -> List[Dict[str, Any]]:
    """
    Seleziona i chunk più rilevanti combinando:
    - similarità coseno sugli embedding
    - match lessicale tra parole chiave della domanda e testo del chunk
    - BOOST se l'utente cita il nome del file nella domanda
    """

    chunks = get_all_chunks()
    if not chunks:
        return []

    question_emb = get_embedding(question)
    if not question_emb:
        return []

    q_tokens = _keyword_tokens(question)
    q_token_set = set(q_tokens)

    scored: List[tuple[float, Dict[str, Any]]] = []
    lexical_candidates: List[Dict[str, Any]] = []

    question_lower = question.lower()

    # 1) primo passaggio: identifichiamo i chunk che contengono almeno una parola chiave
    for rec in chunks:
        text_lower = rec["text"].lower()
        lexical_hits = 0

        # match sulle parole chiave
        for tok in q_token_set:
            if tok in text_lower:
                lexical_hits += 1

        # BOOST esplicito se l'utente cita il nome del file nella domanda
        file_lower = (rec.get("file") or "").lower()
        if file_lower and file_lower in question_lower:
            # incremento forte: equivale a diverse "parole chiave" trovate
            lexical_hits += 3

        if lexical_hits > 0:
            rec["_lex_hits"] = lexical_hits
            lexical_candidates.append(rec)

    # 2) scegliamo il set di chunk da considerare per la similarità coseno
    candidates = lexical_candidates if lexical_candidates else chunks

    # 3) calcoliamo lo score combinato embedding + bonus lessicale
    for rec in candidates:
        emb = rec.get("embedding") or []
        base_score = cosine_similarity(question_emb, emb)

        lex_hits = rec.get("_lex_hits", 0)
        lex_bonus = min(0.3 * lex_hits, 0.9)

        combined_score = base_score + lex_bonus
        scored.append((combined_score, rec))

    # 4) ordiniamo e prendiamo i top_k
    scored.sort(key=lambda x: x[0], reverse=True)

    top_k = max(1, top_k)
    top_chunks = [rec for score, rec in scored[:top_k] if score > 0.0]

    return top_chunks

def get_relevant_chat_chunks(question: str,
                             top_k: int = RAG_TOP_K_DEFAULT) -> List[Dict[str, Any]]:
    """
    Seleziona i chunk più rilevanti tra quelli delle chat salvate.
    Stessa logica combinata embedding + bonus lessicale usata per i documenti.
    """
    chunks = get_all_chat_chunks()
    if not chunks:
        return []

    question_emb = get_embedding(question)
    if not question_emb:
        return []

    q_tokens = _keyword_tokens(question)
    q_token_set = set(q_tokens)

    scored: List[tuple[float, Dict[str, Any]]] = []
    lexical_candidates: List[Dict[str, Any]] = []

    question_lower = question.lower()

    # identifichiamo i chunk che contengono almeno una parola chiave
    for rec in chunks:
        text_lower = rec["text"].lower()
        lexical_hits = 0

        # match sulle parole chiave (numeri, termini, ecc.)
        for tok in q_token_set:
            if tok in text_lower:
                lexical_hits += 1

        # BOOST esplicito se l'utente cita il nome del file nella domanda
        file_lower = (rec.get("file") or "").lower()
        if file_lower and file_lower in question_lower:
            # incremento forte: equivale a diverse "parole chiave" trovate
            lexical_hits += 3

        if lexical_hits > 0:
            # candidato forte: ha almeno una parola chiave o è citato per nome
            rec["_lex_hits"] = lexical_hits
            lexical_candidates.append(rec)


    candidates = lexical_candidates if lexical_candidates else chunks

    for rec in candidates:
        emb = rec.get("embedding") or []
        base_score = cosine_similarity(question_emb, emb)

        lex_hits = rec.get("_lex_hits", 0)
        lex_bonus = min(0.3 * lex_hits, 0.9)

        combined_score = base_score + lex_bonus
        scored.append((combined_score, rec))

    scored.sort(key=lambda x: x[0], reverse=True)

    top_k = max(1, top_k)
    top_chunks = [rec for score, rec in scored[:top_k] if score > 0.0]

    return top_chunks
